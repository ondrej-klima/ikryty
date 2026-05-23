# crud.py
from typing import List, Dict, Any, Optional
from fastapi import HTTPException
from pydantic import BaseModel
from tortoise.exceptions import DoesNotExist, IntegrityError
import logging
from tortoise.transactions import in_transaction
import os
import uuid
import aiofiles
from pathlib import Path
from fastapi import UploadFile
from decimal import Decimal
from io import BytesIO

# Assuming your folder structure is src/crud/crud.py, src/database/models.py etc.
from src.database import models
from src.schemas import schemas
from src.database.models import ShelterLocation

import pyproj
from tortoise.functions import Max, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

"""
Modul pro databázové operace a aplikační logiku.

Tento modul slouží jako prostředník mezi API endpointy (Router) a databázovou vrstvou (Tortoise ORM).
Implementuje základní operace CREATE, READ, UPDATE, DELETE a také logiku pro hodnocení improvizovaných úkrytů.

Hlavní zodpovědnosti modulu:
----------------------------
1.  **Správa Staveb (Buildings):**
    - Vytváření a editace záznamů v Registru staveb (RS).
    - Podpora pro "krokovou" aktualizaci (Step 1-7) dle postupu v metodice.
    - Transformace souřadnic z WGS-84 (GPS) do S-JTSK pomocí knihovny `pyproj` 
      při vytváření nebo úpravě polohy.

2.  **Správa Úkrytů (Shelters):**
    - Evidence improvizovaných úkrytů (RIÚ) vázaných na konkrétní stavby.
    - **Automatické výpočty:** Klíčová funkce `_recalculate_shelter_values` zajišťuje 
      přepočet odvozených parametrů (užitný objem, kapacity N_K/N_KD, koeficienty S_OV, S_O, S_C)
      okamžitě po jakékoli změně vstupních dat.

3.  **Práce se soubory:**
    - Asynchronní nahrávání a mazání příloh (fotodokumentace, schémata, přílohy k S_SD/S_IS).
    - Soubory jsou ukládány na disk do adresářové struktury `uploads/{id}` a cesty jsou
      evidovány v JSON polích v databázi.

4.  **Autorizace:**
    - Interní kontrola oprávnění (`_check_permission`), která zajišťuje, že záznamy
      může editovat/mazat pouze jejich vlastník nebo uživatel s rolí 'supervisor'.

Závislosti:
-----------
- `src.database.models`: Definice DB tabulek (Building, Shelter).
- `src.schemas`: Pydantic modely pro validaci vstupů a výstupů.
- `aiofiles`: Pro asynchronní operace se souborovým systémem.
- `pyproj`: Pro geodetické transformace souřadnic.
"""

logger = logging.getLogger(__name__)

UPLOADS_DIR = Path("uploads")
UPLOADS_DIR.mkdir(exist_ok=True)

BUILDING_EXPORT_FIELDS = [
    "user_id",
    "building_code",
    "name_address",
    "gps_lat",
    "gps_long",
    "owner",
    "administrator",
    "access_restricted",
    "operation_type",
    "object_type",
    "has_underground",
    "has_basement",
    "has_inner_wing",
    "construction_limits",
    "data_source",
    "created_date",
    "responsible_person",
    "risk_area",
    "risk_justification",
    "deficiency",
    "deficiency_justification",
    "wall_material",
    "wall_thickness",
    "s_ok",
    "s_sd",
    "s_sd_attachment",
    "s_is",
    "s_is_attachment",
    "possible_t_upravy_building",
    "last_control_date",
    "control_deficiencies",
    "approver",
    "assessment_status",
    "assessment_date",
    "review_interval",
    "next_review_date",
]

SHELTER_EXPORT_FIELDS = [
    ("user_id", "LOGIN", "Login uživatele"),
    ("building.building_code", "RS_1", "Identifikační kód stavby"),
    ("building.name_address", "RS_2", "Název / adresa stavby"),
    ("building.gps_lat", "RS_3", "GPS souřadnice (šířka, WGS-84)"),
    ("building.gps_long", "RS_3", "GPS souřadnice (délka, WGS-84)"),
    ("shelter_code", None, None),
    ("location", None, None),
    ("schema_path", None, None),
    ("photo_paths", None, None),
    ("area", None, None),
    ("height", None, None),
    ("obstacles_volume", None, None),
    ("usable_volume", None, None),
    ("capacity_short", None, None),
    ("capacity_medium", None, None),
    ("capacity_long", None, None),
    ("expected_persons", None, None),
    ("ventilation", None, None),
    ("emergency_exits", None, None),
    ("power_supply", None, None),
    ("energy_cutoff", None, None),
    ("chuc_type", None, None),
    ("chuc_length", None, None),
    ("chuc_ventilation", None, None),
    ("chuc_walls", None, None),
    ("s_pu", None, None),
    ("s_chuc", None, None),
    ("s_ov", None, None),
    ("distance_to_target", None, None),
    ("s_o", None, None),
    ("s_c", None, None),
    ("iu_class", None, None),
    ("assessment_needed", None, None),
]

BUILDING_FILE_EXPORT_FIELDS = [
    ("user_id", "LOGIN", "Login uživatele"),
    ("building_code", "RS_1", "Identifikační kód stavby"),
    ("name_address", "RS_2", "Název / adresa stavby"),
    ("gps_lat", "RS_3", "GPS souřadnice (šířka, WGS-84)"),
    ("gps_long", "RS_3", "GPS souřadnice (délka, WGS-84)"),
    ("attachment_type", "SOUBOR", "Typ přílohy"),
    ("filename", "SOUBOR", "Název souboru"),
]

SHELTER_FILE_EXPORT_FIELDS = [
    ("user_id", "LOGIN", "Login uživatele"),
    ("building_code", "RS_1", "Identifikační kód stavby"),
    ("building_name_address", "RS_2", "Název / adresa stavby"),
    ("gps_lat", "RS_3", "GPS souřadnice (šířka, WGS-84)"),
    ("gps_long", "RS_3", "GPS souřadnice (délka, WGS-84)"),
    ("shelter_code", "RIÚ_4", "Identifikační kód IÚ"),
    ("attachment_type", "SOUBOR", "Typ přílohy"),
    ("filename", "SOUBOR", "Název souboru"),
]


# ==============================================================================
# AUTHORIZATION & HELPER FUNCTIONS
# ==============================================================================

def _check_permission(db_object_user_id: str, current_user: Dict[str, Any]):
    """
       Ověří, zda má aktuální uživatel oprávnění přistupovat k objektu nebo jej modifikovat.

       Přístup je povolen, pokud je uživatel vlastníkem objektu (shoda ID)
       nebo má roli 'supervisor'.

       Args:
           db_object_user_id (str): ID uživatele (username), který vlastní objekt v databázi.
           current_user (Dict[str, Any]): Slovník s informacemi o přihlášeném uživateli (z JWT).

       Raises:
           HTTPException: 403 Forbidden, pokud uživatel nemá oprávnění.
       """
    username = current_user.get('preferred_username')
    roles = current_user.get("realm_access", {}).get("roles", [])
    if not (db_object_user_id == username or 'supervisor' in roles):
        raise HTTPException(status_code=403, detail="Not authorized to perform this action")


def _is_supervisor(current_user: Dict[str, Any]) -> bool:
    return 'supervisor' in current_user.get("realm_access", {}).get("roles", [])


def _get_visible_buildings_query(current_user: Dict[str, Any], building_ids: Optional[List[int]] = None):
    if _is_supervisor(current_user):
        query = models.Building.all()
    else:
        query = models.Building.filter(user_id=current_user.get('preferred_username'))

    if building_ids is not None:
        query = query.filter(id__in=building_ids)

    return query


def _parse_field_description(description: str) -> tuple[str, str]:
    if not description:
        return "", ""

    abbreviation, separator, caption = description.partition(":")
    if not separator:
        return "", description.strip()
    return abbreviation.strip(), caption.strip()


def _format_export_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, bool):
        return "ANO" if value else "NE"
    if hasattr(value, "value"):
        return value.value
    if value is None:
        return ""
    return value


def _resolve_attr(source: Any, path: str) -> Any:
    value = source
    for part in path.split("."):
        value = getattr(value, part, None)
        if value is None:
            return None
    return value


def _set_sheet_headers(worksheet, header_rows: List[tuple[str, str]]):
    for column_index, (caption, abbreviation) in enumerate(header_rows, start=1):
        caption_cell = worksheet.cell(row=1, column=column_index, value=caption)
        abbreviation_cell = worksheet.cell(row=2, column=column_index, value=abbreviation)
        caption_cell.font = Font(bold=True)
        abbreviation_cell.font = Font(bold=True)
        caption_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        abbreviation_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(worksheet):
    for column in worksheet.columns:
        values = [cell.value for cell in column if cell.value is not None]
        if not values:
            continue
        width = min(max(len(str(value)) for value in values) + 2, 60)
        worksheet.column_dimensions[column[0].column_letter].width = width


def _has_valid_coordinates(latitude: Any, longitude: Any) -> bool:
    return latitude not in (None, 0, 0.0, "") and longitude not in (None, 0, 0.0, "")


def _extract_filename(path_value: str) -> str:
    return Path(str(path_value)).name


def _append_row(worksheet, values: List[Any]):
    worksheet.append([_format_export_value(value) for value in values])


def _build_export_workbook(buildings: List[models.Building]) -> bytes:
    workbook = Workbook()
    buildings_sheet = workbook.active
    buildings_sheet.title = "Registr staveb"
    shelters_sheet = workbook.create_sheet("Registr improvizovaných úkrytů")
    building_files_sheet = workbook.create_sheet("Soubory staveb")
    shelter_files_sheet = workbook.create_sheet("Soubory úkrytů")

    building_headers = []
    for field_name in BUILDING_EXPORT_FIELDS:
        abbreviation, caption = _parse_field_description(models.Building._meta.fields_map[field_name].description)
        building_headers.append((caption, abbreviation))
    building_headers[0] = ("Login uživatele", "LOGIN")
    _set_sheet_headers(buildings_sheet, building_headers)

    shelter_headers = []
    for field_name, abbreviation, caption in SHELTER_EXPORT_FIELDS:
        if abbreviation is None or caption is None:
            abbreviation, caption = _parse_field_description(models.Shelter._meta.fields_map[field_name].description)
        shelter_headers.append((caption, abbreviation))
    _set_sheet_headers(shelters_sheet, shelter_headers)

    _set_sheet_headers(building_files_sheet, [(caption, abbreviation) for _, abbreviation, caption in BUILDING_FILE_EXPORT_FIELDS])
    _set_sheet_headers(shelter_files_sheet, [(caption, abbreviation) for _, abbreviation, caption in SHELTER_FILE_EXPORT_FIELDS])

    for building in buildings:
        if not _has_valid_coordinates(building.gps_lat, building.gps_long):
            continue

        _append_row(
            buildings_sheet,
            [getattr(building, field_name, None) for field_name in BUILDING_EXPORT_FIELDS],
        )

        for attachment_type, field_name in (("Přílohy k S_SD", "s_sd_attachment"), ("Přílohy k S_IS", "s_is_attachment")):
            for attachment_path in getattr(building, field_name, None) or []:
                _append_row(
                    building_files_sheet,
                    [
                        building.user_id,
                        building.building_code,
                        building.name_address,
                        building.gps_lat,
                        building.gps_long,
                        attachment_type,
                        _extract_filename(attachment_path),
                    ],
                )

        for shelter in building.shelters:
            shelter_row = []
            for field_name, _, _ in SHELTER_EXPORT_FIELDS:
                source = building if field_name.startswith("building.") else shelter
                path = field_name.removeprefix("building.")
                shelter_row.append(_resolve_attr(source, path))
            _append_row(shelters_sheet, shelter_row)

            for attachment_type, field_name in (("Schéma prostoru", "schema_path"), ("Fotodokumentace", "photo_paths")):
                for attachment_path in getattr(shelter, field_name, None) or []:
                    _append_row(
                        shelter_files_sheet,
                        [
                            shelter.user_id,
                            building.building_code,
                            building.name_address,
                            building.gps_lat,
                            building.gps_long,
                            shelter.shelter_code,
                            attachment_type,
                            _extract_filename(attachment_path),
                        ],
                    )

    buildings_sheet.freeze_panes = "A3"
    shelters_sheet.freeze_panes = "A3"
    building_files_sheet.freeze_panes = "A3"
    shelter_files_sheet.freeze_panes = "A3"
    _autosize_columns(buildings_sheet)
    _autosize_columns(shelters_sheet)
    _autosize_columns(building_files_sheet)
    _autosize_columns(shelter_files_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _recalculate_shelter_values(shelter: models.Shelter):
    """
    Přepočítá všechny závislé (vypočítané) hodnoty pro daný úkryt a uloží změny do DB.
    Tato funkce musí být zavolána po jakékoli změně vstupních dat úkrytu nebo budovy.

    Provádí následující výpočty:
    1. Užitný objem a kapacity (krátkodobá, střednědobá, dlouhodobá).
    2. Koeficienty polohy (S_PU) a CHÚC (S_CHUC).
    3. Koeficient ochranných vlastností stavby (S_OV).
    4. Koeficient vzdálenosti (S_O).
    5. Výsledný koeficient ochrany (S_C) a klasifikace úkrytu (IÚ/KIÚ).

    Args:
        shelter (models.Shelter): Instance modelu úkrytu.
    """

    building = await shelter.building

    # A3: Capacity Calculation
    if shelter.area is not None and shelter.height is not None:
        v_tp = shelter.obstacles_volume or 0
        v = (shelter.area * shelter.height) - v_tp
        if v > 0:
            shelter.usable_volume = round(v, 2)
            shelter.capacity_short = v // 3
            shelter.capacity_medium = v // 5
            shelter.capacity_long = v // 8
        else:
            shelter.usable_volume = shelter.capacity_short = shelter.capacity_medium = shelter.capacity_long = None

    if shelter.location is not None:
        shelter.s_pu = 0 if shelter.location == ShelterLocation.INNER_WING else 1

    if shelter.is_chuc is not None:
        shelter.s_chuc = 1 if shelter.is_chuc else 0

    # A4: S_OV Calculation
    if all(v is not None for v in [building.s_ok, building.s_sd, building.s_is, shelter.s_pu, shelter.s_chuc]):
        s_ov = (0.4 * building.s_ok) + (0.2 * building.s_sd) + (0.1 * building.s_is) + \
               (0.2 * shelter.s_pu) + (0.1 * shelter.s_chuc)
        shelter.s_ov = round(s_ov, 3)
    else:
        shelter.s_ov = None

    # A5: S_O Calculation
    if shelter.distance_to_target is not None:
        if shelter.distance_to_target < 100:
            shelter.s_o = 3
        elif shelter.distance_to_target <= 500:
            shelter.s_o = 2
        else:
            shelter.s_o = 1
    else:
        shelter.s_o = None

    # A6: S_C and Classification Calculation
    if all(v is not None and v > 0 for v in [shelter.s_ov, shelter.capacity_short, shelter.s_o]) and \
            all(v is not None for v in [building.s_sd, building.s_is]):
        #s_c = (1000 * shelter.s_ov) / (2 * shelter.capacity_short * shelter.s_o)

        s_ov = shelter.s_ov
        n_k = shelter.capacity_short
        s_o = shelter.s_o

        # New formula: S_C = (10 * S_OV) / (((2 * N_K) / 100)**(1/3) * S_O)
        numerator = 10 * s_ov
        # Calculate the cube root term safely
        cube_root_term = ((2 * n_k) / 100) ** (1 / 3)
        denominator = cube_root_term * s_o

        if denominator > 0:
            s_c = numerator / denominator
            shelter.s_c = round(s_c, 3)
        else:
            # Avoid division by zero
            shelter.s_c = None

        #shelter.s_c = round(s_c, 3)

        if s_c >= 5:
            shelter.iu_class = 'IÚ-I' if building.s_sd == 1 and building.s_is == 1 else 'KIÚ-I'
        elif s_c >= 3:
            shelter.iu_class = 'IÚ-II' if building.s_sd == 1 and building.s_is == 1 else 'KIÚ-II'
        elif s_c >= 1:
            shelter.iu_class = 'KIÚ-III'
        else:
            shelter.iu_class = 'KIÚ-N – nevhodné'

        shelter.assessment_needed = 'ANO' if 'KIÚ' in (shelter.iu_class or '') and shelter.iu_class != 'KIÚ-N – nevhodné'  else 'NE'
    else:
        shelter.s_c = shelter.iu_class = shelter.assessment_needed = None

    await shelter.save()


# ==============================================================================
# BUILDING CRUD OPERATIONS
# ==============================================================================

async def create_building(building_in: schemas.BuildingInSchema,
                          current_user: Dict[str, Any]) -> schemas.BuildingSimpleOutSchema:
    """
    Vytvoří novou budovu v databázi.

    Automaticky transformuje zadané GPS souřadnice (WGS84) na souřadný systém S-JTSK.

    Args:
        building_in (schemas.BuildingInSchema): Vstupní data pro vytvoření budovy.
        current_user (Dict[str, Any]): Data přihlášeného uživatele.

    Returns:
        schemas.BuildingSimpleOutSchema: Vytvořená budova (zjednodušené schéma).

    Raises:
        HTTPException: 409 Conflict při duplicitním kódu budovy, 500 při jiné chybě.
    """
    try:
        crs_wgs84 = pyproj.CRS("EPSG:4326")
        crs_s_jtsk = pyproj.CRS("EPSG:5514")

        transformer = pyproj.Transformer.from_crs(crs_wgs84, crs_s_jtsk, always_xy=True)

        x, y = transformer.transform(building_in.gps_long, building_in.gps_lat)
        building_in.s_jtsk_x = x
        building_in.s_jtsk_y = y

        building_obj = await models.Building.create(
            **building_in.dict(),
            user_id=current_user.get('preferred_username')
        )
        return await schemas.BuildingSimpleOutSchema.from_tortoise_orm(building_obj)
    except IntegrityError as e:
        logger.error(f"Database integrity error during building creation: {e} | Cause: {e.__cause__}")
        original_error_msg = str(e.__cause__).lower()
        if "duplicate entry" in original_error_msg and "building_code" in original_error_msg:
            detail_message = f"Stavba s kódem '{building_in.building_code}' již existuje."
        else:
            detail_message = "Chyba databázové integrity při vytváření stavby."
        raise HTTPException(status_code=409, detail=detail_message)
    except Exception as e:
        logger.error(f"An unexpected error occurred during building creation: {e}")
        #print(f"An unexpected error occurred during building creation: {e}")
        raise HTTPException(status_code=500, detail="Internal Server Error")


async def get_all_buildings() -> List[schemas.BuildingOutSchema]:
    """
    Získá seznam všech budov uložených v databázi.
    Zahrnuje i načtení souvisejících úkrytů (prefetch).

    Returns:
        List[schemas.BuildingOutSchema]: Seznam budov.
    """
    buildings = await models.Building.all().prefetch_related("shelters")
    return [await schemas.BuildingOutSchema.from_tortoise_orm(b) for b in buildings]


async def get_building(building_id: int, current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """
    Získá detail konkrétní budovy podle ID.

    Args:
        building_id (int): ID budovy.
        current_user (Dict[str, Any]): Data přihlášeného uživatele (pro kontrolu oprávnění).

    Returns:
        schemas.BuildingOutSchema: Detail budovy.

    Raises:
        HTTPException: 404 Not Found, pokud budova neexistuje.
    """
    building = await models.Building.get_or_none(id=building_id).prefetch_related("shelters")
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID {building_id} nenalezena.")
    _check_permission(building.user_id, current_user)
    return await schemas.BuildingOutSchema.from_tortoise_orm(building)


async def delete_building(building_id: int, current_user: Dict[str, Any]) -> schemas.Status:
    """
    Smaže budovu a všechny k ní přidružené přístřešky.
    Operace probíhá v rámci databázové transakce pro zajištění integrity dat.

    Args:
        building_id (int): ID budovy ke smazání.
        current_user (Dict[str, Any]): Data přihlášeného uživatele.

    Returns:
        schemas.Status: Potvrzení o smazání.

    Raises:
        HTTPException: 404 Not Found, pokud budova neexistuje.
    """
    # 1. Získání objektu budovy
    # Použijeme .prefetch_related('shelters'), což může být mírně efektivnější,
    # i když pro samotné mazání to není striktně nutné.
    building = await models.Building.get_or_none(id=building_id)
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID {building_id} nenalezena.")

    # 3. Provedení operací v rámci transakce
    async with in_transaction():
        # 3a. Smazání všech souvisejících přístřešků
        # Přistoupíme k nim přes related_name 'shelters' a na querysetu zavoláme .delete()
        await building.shelters.all().delete()

        # 3b. Smazání samotné budovy
        await building.delete()

    return schemas.Status(message=f"Smazána stavba s ID {building_id} a všechny související přístřešky.")


async def _update_building_and_recalculate(building_id: int, data_in: BaseModel,
                                           current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """
    Interní pomocná funkce pro aktualizaci budovy a následný přepočet.

    Kroky:
    1. Aktualizuje data budovy.
    2. Pokud se změnily GPS souřadnice, přepočítá S-JTSK.
    3. Spustí přepočet (`_recalculate_shelter_values`) pro VŠECHNY úkryty v budově,
       protože parametry budovy ovlivňují parametry úkrytů.

    Args:
        building_id (int): ID budovy.
        data_in (BaseModel): Pydantic model s novými daty.
        current_user (Dict[str, Any]): Data přihlášeného uživatele.

    Returns:
        schemas.BuildingOutSchema: Aktualizovaná budova.
    """
    building = await models.Building.get_or_none(id=building_id)
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID {building_id} nenalezena.")
    _check_permission(building.user_id, current_user)

    if hasattr(data_in, 'gps_lat') and hasattr(data_in, 'gps_long'):
        if data_in.gps_lat is not None and data_in.gps_long is not None:
            crs_wgs84 = pyproj.CRS("EPSG:4326")
            crs_s_jtsk = pyproj.CRS("EPSG:5514")

            transformer = pyproj.Transformer.from_crs(crs_wgs84, crs_s_jtsk, always_xy=True)

            x, y = transformer.transform(data_in.gps_long, data_in.gps_lat)
            data_in.s_jtsk_x = x
            data_in.s_jtsk_y = y

    update_data = data_in.dict(exclude_unset=True)
    update_data['user_id'] = current_user.get('preferred_username')
    await building.update_from_dict(update_data).save()

    await building.fetch_related("shelters")
    for shelter in building.shelters:
        await _recalculate_shelter_values(shelter)

    # Reload the full building with updated shelters to return the freshest data
    return await get_building(building_id, current_user)


async def update_building_step1(building_id: int, data_in: schemas.BuildingStep1UpdateSchema,
                                current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Aktualizuje budovu - Krok 1 (Základní údaje)."""
    return await _update_building_and_recalculate(building_id, data_in, current_user)


async def update_building_step2(building_id: int, data_in: schemas.BuildingStep2UpdateSchema,
                                current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Aktualizuje budovu - Krok 2 (Konstrukční údaje)."""
    return await _update_building_and_recalculate(building_id, data_in, current_user)


async def update_building_step3(building_id: int, data_in: schemas.BuildingStep3UpdateSchema,
                                current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Aktualizuje budovu - Krok 3 (Údaje o okolí)."""
    return await _update_building_and_recalculate(building_id, data_in, current_user)


async def update_building_step4(building_id: int, data_in: schemas.BuildingStep4UpdateSchema,
                                current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Aktualizuje budovu - Krok 4 (Doplňkové údaje)."""
    return await _update_building_and_recalculate(building_id, data_in, current_user)


async def update_building_step7(building_id: int, data_in: schemas.BuildingStep7UpdateSchema,
                                current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Aktualizuje budovu - Krok 7 (Závěrečné údaje)."""
    return await _update_building_and_recalculate(building_id, data_in, current_user)


# ==============================================================================
# SHELTER CRUD OPERATIONS
# ==============================================================================

async def create_shelter(shelter_in: schemas.ShelterInSchema, current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """
    Vytvoří nový úkryt v rámci existující budovy.

    Po vytvoření automaticky provede výpočet odvozených hodnot.

    Args:
        shelter_in (schemas.ShelterInSchema): Vstupní data úkrytu (včetně ID budovy).
        current_user (Dict[str, Any]): Data přihlášeného uživatele.

    Returns:
        schemas.ShelterOutSchema: Vytvořený úkryt.

    Raises:
        HTTPException: 404 pokud budova neexistuje, 409 pokud úkryt s kódem již existuje.
    """
    building = await models.Building.get_or_none(id=shelter_in.building_id)
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID '{shelter_in.building_id}' neexistuje.")
    try:
        shelter_data = shelter_in.dict()
        shelter_data['user_id'] = current_user.get('preferred_username')
        shelter_obj = await models.Shelter.create(**shelter_data, building=building)
        await _recalculate_shelter_values(shelter_obj)
        return await schemas.ShelterOutSchema.from_tortoise_orm(shelter_obj)
    except IntegrityError:
        raise HTTPException(status_code=409,
                            detail=f"Úkryt s kódem '{shelter_in.shelter_code}' již v této stavbě existuje.")


async def delete_shelter(shelter_id: int, current_user: Dict[str, Any]) -> schemas.Status:
    """
    Smaže konkrétní úkryt.

    Args:
        shelter_id (int): ID úkrytu.
        current_user (Dict[str, Any]): Data přihlášeného uživatele.

    Returns:
        schemas.Status: Potvrzení o smazání.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)
    await shelter.delete()
    return schemas.Status(message=f"Smazán úkryt s ID {shelter_id}")


async def _update_shelter_and_recalculate(shelter_id: int, data_in: BaseModel,
                                          current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """
    Interní pomocná funkce pro aktualizaci dat úkrytu a přepočet hodnot.

    Args:
        shelter_id (int): ID úkrytu.
        data_in (BaseModel): Data pro aktualizaci.
        current_user (Dict[str, Any]): Přihlášený uživatel.

    Returns:
        schemas.ShelterOutSchema: Aktualizovaný úkryt.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)

    update_data = data_in.dict(exclude_unset=True)
    update_data['user_id'] = current_user.get('preferred_username')
    await shelter.update_from_dict(update_data).save()

    await _recalculate_shelter_values(shelter)

    return await schemas.ShelterOutSchema.from_tortoise_orm(shelter)


async def update_shelter_step3(shelter_id: int, data_in: schemas.ShelterStep3UpdateSchema,
                               current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """Aktualizuje úkryt - Krok 3 (Parametry úkrytu)."""
    return await _update_shelter_and_recalculate(shelter_id, data_in, current_user)


async def update_shelter_step4(shelter_id: int, data_in: schemas.ShelterStep4UpdateSchema,
                               current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """Aktualizuje úkryt - Krok 4 (Specifikace)."""
    return await _update_shelter_and_recalculate(shelter_id, data_in, current_user)


async def update_shelter_step5(shelter_id: int, data_in: schemas.ShelterStep5UpdateSchema,
                               current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """Aktualizuje úkryt - Krok 5 (Další vlastnosti)."""
    return await _update_shelter_and_recalculate(shelter_id, data_in, current_user)


async def update_shelter_step6(shelter_id: int, data_in: schemas.ShelterStep6UpdateSchema,
                               current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """
    Aktualizuje úkryt - Krok 6.
    Pouze ukládá data, neprovádí složitý přepočet hodnot.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)
    update_data = data_in.dict(exclude_unset=True)
    update_data['user_id'] = current_user.get('preferred_username')
    await shelter.update_from_dict(update_data).save()
    return await schemas.ShelterOutSchema.from_tortoise_orm(shelter)


# ==============================================================================
# PHOTO UPLOAD/DELETE OPERATIONS
# ==============================================================================

async def upload_photos_for_shelter(shelter_id: int, files: List[UploadFile],
                                    current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """
    Nahraje fotografie ke konkrétnímu úkrytu.

    Soubory jsou uloženy na disk do složky `uploads/{shelter_id}` a jejich cesty
    jsou přidány do seznamu v databázi.

    Args:
        shelter_id (int): ID úkrytu.
        files (List[UploadFile]): Seznam nahrávaných souborů.
        current_user (Dict[str, Any]): Přihlášený uživatel.

    Returns:
        schemas.ShelterOutSchema: Aktualizovaný objekt úkrytu.

    Raises:
        HTTPException: 500 při chybě zápisu na disk.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)

    # Ensure the photo_paths list exists
    if shelter.photo_paths is None:
        shelter.photo_paths = []

    # Create a unique subdirectory for this shelter's uploads
    shelter_upload_dir = UPLOADS_DIR / str(shelter.id)
    shelter_upload_dir.mkdir(exist_ok=True)

    for file in files:
        # Generate a safe, unique filename
        file_extension = Path(file.filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = shelter_upload_dir / unique_filename

        # Save the file asynchronously
        try:
            async with aiofiles.open(file_path, 'wb') as out_file:
                content = await file.read()
                await out_file.write(content)

            # Store the relative path in the database
            relative_path = os.path.join(str(shelter.id), unique_filename)
            shelter.photo_paths.append(relative_path)

        except Exception as e:
            logger.error(f"Could not save file {file.filename}: {e}")
            raise HTTPException(status_code=500, detail="Chyba při nahrávání souboru.")

    shelter.user_id = current_user.get('preferred_username')  # Update the last editor
    await shelter.save()
    return await schemas.ShelterOutSchema.from_tortoise_orm(shelter)


async def delete_photo_for_shelter(shelter_id: int, filename: str, current_user: Dict[str, Any]) -> schemas.Status:
    """
    Smaže fotografii úkrytu z databáze i z disku.

    Args:
        shelter_id (int): ID úkrytu.
        filename (str): Název souboru (pouze název, ne celá cesta).
        current_user (Dict[str, Any]): Přihlášený uživatel.

    Returns:
        schemas.Status: Potvrzení o smazání.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)

    # Construct the relative path as it is stored in the DB
    relative_path = os.path.join(str(shelter.id), filename)

    if shelter.photo_paths and relative_path in shelter.photo_paths:
        # 1. Remove from database
        shelter.photo_paths.remove(relative_path)
        shelter.user_id = current_user.get('preferred_username')
        await shelter.save()

        # 2. Delete from disk
        full_file_path = UPLOADS_DIR / relative_path
        try:
            os.remove(full_file_path)
        except FileNotFoundError:
            logger.warning(f"File not found on disk but was removed from DB: {full_file_path}")
        except Exception as e:
            logger.error(f"Error deleting file from disk: {e}")
            # Don't raise an error to the user, as the DB record is already fixed.

        return schemas.Status(message=f"Fotografie '{filename}' byla smazána.")
    else:
        raise HTTPException(status_code=404, detail=f"Fotografie '{filename}' nebyla u tohoto úkrytu nalezena.")


# ==============================================================================
# SCHEMA UPLOAD/DELETE OPERATIONS
# ==============================================================================

async def upload_shelter_schemas(shelter_id: int, files: List[UploadFile],
                                 current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """Nahraje soubory schémat pro daný úkryt (využívá generickou funkci)."""
    return await _upload_files_for_shelter(shelter_id, files, "schema_path", current_user)


async def delete_shelter_schema(shelter_id: int, filename: str, current_user: Dict[str, Any]) -> schemas.Status:
    """Smaže soubor schématu pro daný úkryt (využívá generickou funkci)."""
    return await _delete_file_for_shelter(shelter_id, filename, "schema_path", current_user)


# It's good practice to create generic helpers to stay DRY (Don't Repeat Yourself)
# These helpers can serve both photos and schemas

async def _upload_files_for_shelter(shelter_id: int, files: List[UploadFile], field_name: str,
                                    current_user: Dict[str, Any]) -> schemas.ShelterOutSchema:
    """
    Generická pomocná funkce pro nahrávání souborů k úkrytu do dynamického pole.

    Args:
        shelter_id (int): ID úkrytu.
        files (List[UploadFile]): Seznam souborů.
        field_name (str): Název atributu modelu (např. 'schema_path' nebo 'photo_paths').
        current_user (Dict[str, Any]): Uživatel.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)

    if getattr(shelter, field_name) is None:
        setattr(shelter, field_name, [])

    shelter_upload_dir = UPLOADS_DIR / str(shelter.id)
    shelter_upload_dir.mkdir(exist_ok=True)

    for file in files:
        file_extension = Path(file.filename).suffix
        unique_filename = f"{field_name}_{uuid.uuid4()}{file_extension}"
        file_path = shelter_upload_dir / unique_filename

        try:
            async with aiofiles.open(file_path, 'wb') as out_file:
                content = await file.read()
                await out_file.write(content)

            relative_path = os.path.join(str(shelter.id), unique_filename)
            getattr(shelter, field_name).append(relative_path)
        except Exception as e:
            # handle error
            pass

    shelter.user_id = current_user.get('preferred_username')
    await shelter.save()
    return await schemas.ShelterOutSchema.from_tortoise_orm(shelter)


async def _delete_file_for_shelter(shelter_id: int, filename: str, field_name: str,
                                   current_user: Dict[str, Any]) -> schemas.Status:
    """
    Generická pomocná funkce pro mazání souborů úkrytu.
    """
    shelter = await models.Shelter.get_or_none(id=shelter_id)
    if not shelter:
        raise HTTPException(status_code=404, detail=f"Úkryt s ID {shelter_id} nenalezen.")
    _check_permission(shelter.user_id, current_user)

    file_list = getattr(shelter, field_name)
    target_path = next((path for path in file_list if path.endswith(filename)), None)

    if target_path:
        file_list.remove(target_path)
        shelter.user_id = current_user.get('preferred_username')
        await shelter.save()

        full_file_path = UPLOADS_DIR / target_path
        try:
            os.remove(full_file_path)
        except FileNotFoundError:
            pass  # Ignore if file is already gone

        return schemas.Status(message=f"Soubor '{filename}' byl smazán.")
    else:
        raise HTTPException(status_code=404, detail=f"Soubor '{filename}' nebyl u tohoto úkrytu nalezen.")

# ==============================================================================
# BUILDING FILE UPLOAD/DELETE OPERATIONS
# ==============================================================================

async def _upload_building_attachments(building_id: int, files: List[UploadFile], field_name: str,
                                       current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """
    Generická pomocná funkce pro nahrání jedné nebo více příloh k budově.

    Ukládá soubory do složky `uploads/building_{id}`.

    Args:
        building_id (int): ID budovy.
        files (List[UploadFile]): Seznam souborů.
        field_name (str): Název pole v DB modelu (např. 's_sd_attachment').
        current_user (Dict[str, Any]): Uživatel.

    Returns:
        schemas.BuildingOutSchema: Aktualizovaná budova.
    """
    building = await models.Building.get_or_none(id=building_id)
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID {building_id} nenalezena.")
    _check_permission(building.user_id, current_user)

    # Ensure the attachment list exists
    if getattr(building, field_name) is None:
        setattr(building, field_name, [])

    building_upload_dir = UPLOADS_DIR / f"building_{building.id}"
    building_upload_dir.mkdir(exist_ok=True)

    for file in files:
        file_extension = Path(file.filename).suffix
        unique_filename = f"{field_name}_{uuid.uuid4()}{file_extension}"
        file_path = building_upload_dir / unique_filename

        try:
            async with aiofiles.open(file_path, 'wb') as out_file:
                content = await file.read()
                await out_file.write(content)

            relative_path = os.path.join(f"building_{building.id}", unique_filename)
            getattr(building, field_name).append(relative_path)

        except Exception as e:
            logger.error(f"Could not save attachment for building {building_id}: {e}")
            raise HTTPException(status_code=500, detail="Chyba při nahrávání souboru.")

    building.user_id = current_user.get('preferred_username')
    await building.save()
    return await get_building(building_id, current_user)


async def _delete_building_attachment(building_id: int, filename: str, field_name: str,
                                      current_user: Dict[str, Any]) -> schemas.Status:
    """
    Generická pomocná funkce pro smazání konkrétní přílohy budovy.
    """
    building = await models.Building.get_or_none(id=building_id)
    if not building:
        raise HTTPException(status_code=404, detail=f"Stavba s ID {building_id} nenalezena.")
    _check_permission(building.user_id, current_user)

    attachment_list = getattr(building, field_name)

    # Find the full path that ends with the given filename
    target_path = None
    if attachment_list:
        for path in attachment_list:
            if path.endswith(filename):
                target_path = path
                break

    if target_path:
        # 1. Remove from database list
        attachment_list.remove(target_path)
        building.user_id = current_user.get('preferred_username')
        await building.save()

        # 2. Delete from disk
        full_file_path = UPLOADS_DIR / target_path
        try:
            os.remove(full_file_path)
        except FileNotFoundError:
            logger.warning(f"File not found on disk but was removed from DB: {full_file_path}")

        return schemas.Status(message=f"Příloha '{filename}' byla smazána.")
    else:
        raise HTTPException(status_code=404, detail=f"Příloha '{filename}' nebyla u tohoto pole nalezena.")


# Specific functions now call the plural helpers
async def upload_building_ssd_attachments(building_id: int, files: List[UploadFile],
                                          current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Nahraje přílohy k parametru S_SD budovy."""
    return await _upload_building_attachments(building_id, files, "s_sd_attachment", current_user)


async def delete_building_ssd_attachment(building_id: int, filename: str,
                                         current_user: Dict[str, Any]) -> schemas.Status:
    """Smaže přílohu parametru S_SD budovy."""
    return await _delete_building_attachment(building_id, filename, "s_sd_attachment", current_user)


async def upload_building_sis_attachments(building_id: int, files: List[UploadFile],
                                          current_user: Dict[str, Any]) -> schemas.BuildingOutSchema:
    """Nahraje přílohy k parametru S_IS budovy."""
    return await _upload_building_attachments(building_id, files, "s_is_attachment", current_user)


async def delete_building_sis_attachment(building_id: int, filename: str,
                                         current_user: Dict[str, Any]) -> schemas.Status:
    """Smaže přílohu parametru S_IS budovy."""
    return await _delete_building_attachment(building_id, filename, "s_is_attachment", current_user)


async def get_buildings_summary(current_user: Dict[str, Any]) -> List[schemas.BuildingSummarySchema]:
    """
    Získá přehledový seznam budov pro zobrazení na mapě nebo v tabulce.

    Pro každou budovu provádí agregaci (anotaci), kdy nalezne maximální hodnotu
    koeficientu ochranných vlastností (S_C) ze všech jejích úkrytů (`max_s_c`).

    Args:
        current_user (Dict[str, Any]): Přihlášený uživatel.

    Returns:
        List[schemas.BuildingSummarySchema]: Seznam souhrnných dat budov.

    Raises:
        HTTPException: 500 při chybě validace dat (Pydantic).
    """
    print("\n--- RUNNING get_buildings_summary ---")

    # 1. The database query is correct.
    buildings_data = await _get_visible_buildings_query(current_user).annotate(
        max_s_c=Max("shelters__s_c"),
        total_n_k=Sum("shelters__capacity_short"),
        total_n_ks=Sum("shelters__capacity_medium"),
        total_n_kd=Sum("shelters__capacity_long"),
    ).values(
        "id", "building_code", "user_id", "name_address", "gps_lat", "gps_long", "max_s_c", "total_n_k", "total_n_ks", "total_n_kd"
    )

    if not buildings_data:
        print("--- No buildings found, returning empty list. ---")
        return []

    # --- DEBUGGING: Print the raw data and its types ---
    print("\n--- RAW DATA AND TYPES FROM FIRST RECORD ---")
    first_record = buildings_data[0]
    for key, value in first_record.items():
        print(f"  - Key: '{key}', Value: {repr(value)}, Type: {type(value)}")
    print("-" * 40)
    # --- END DEBUGGING ---

    results = []
    for building_dict in buildings_data:
        # 2. DEFENSIVE TYPE CONVERSION
        max_s_c = building_dict.get("max_s_c")
        if isinstance(max_s_c, Decimal):
            max_s_c = float(max_s_c)

        gps_lat = building_dict.get("gps_lat")
        if gps_lat is not None:
            gps_lat = float(gps_lat)

        gps_long = building_dict.get("gps_long")
        if gps_long is not None:
            gps_long = float(gps_long)

        total_n_k = int(building_dict.get("total_n_k") or 0)
        total_n_ks = int(building_dict.get("total_n_ks") or 0)
        total_n_kd = int(building_dict.get("total_n_kd") or 0)

        # 3. Manually construct the Pydantic object
        try:
            summary_obj = schemas.BuildingSummarySchema(
                id=int(building_dict["id"]),
                building_code=str(building_dict["building_code"]),
                user_id=str(building_dict["user_id"]),
                name_address=str(building_dict["name_address"]),
                gps_lat=gps_lat,
                gps_long=gps_long,
                max_s_c=max_s_c,
                total_n_k=total_n_k,
                total_n_ks=total_n_ks,
                total_n_kd=total_n_kd,
            )
            results.append(summary_obj)
        except Exception as pydantic_error:
            print(f"!!! PYDANTIC VALIDATION FAILED for dict: {building_dict} !!!")
            print(f"Pydantic Error: {pydantic_error}")
            raise HTTPException(status_code=500, detail="Internal server error during data processing.")

    return results


async def export_visible_buildings_and_shelters(current_user: Dict[str, Any], building_ids: Optional[List[int]] = None) -> bytes:
    buildings = await _get_visible_buildings_query(current_user, building_ids=building_ids).prefetch_related("shelters")
    return _build_export_workbook(buildings)
